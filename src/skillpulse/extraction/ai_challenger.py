"""Repair external AI annotations into a canonical, non-human challenger artifact."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from .agreement import evaluate_annotation_agreement
from .annotation_quality import (
    ALLOWED_ARRANGEMENT,
    ALLOWED_EDUCATION,
    ALLOWED_SENIORITY,
    REQUIRED_COLUMNS,
    validate_annotation_frame,
)
from .engine import EntityExtractor
from .review_batch import GOLD_COLUMNS, _source_key, _text_sha256, write_csv_atomic

AI_CHALLENGER_COLUMNS = (
    "source_row",
    "text",
    "text_sha256",
    *GOLD_COLUMNS,
    "review_status",
    "annotator",
    "notes",
)
SET_FIELDS = {
    "gold_technical_skills": "technical_skill",
    "gold_tools": "tool",
    "gold_soft_skills": "soft_skill",
}


@dataclass(frozen=True)
class AIRepairResult:
    annotations: pd.DataFrame
    report: dict[str, Any]


def _clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _split(value: Any) -> list[str]:
    return [label.strip() for label in _clean(value).split("|") if label.strip()]


def _normalized_years(value: Any) -> str:
    cleaned = _clean(value)
    if not cleaned:
        return ""
    try:
        numeric = float(cleaned)
    except ValueError:
        return ""
    return f"{numeric:g}" if numeric >= 0 else ""


def _require_columns(frame: pd.DataFrame, required: set[str], name: str) -> None:
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"Missing {name} columns: {', '.join(missing)}")


def read_annotation_workbook(path: Path, sheet_name: str = "Gold_Annotations") -> pd.DataFrame:
    """Read an annotation worksheet without evaluating formulas or mutating the file."""
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet {sheet_name!r} not found in {path}")
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Worksheet {sheet_name!r} is empty")
    headers = [_clean(value) for value in rows[0]]
    if not all(headers):
        raise ValueError("Annotation worksheet contains blank header cells")
    return pd.DataFrame(rows[1:], columns=headers).fillna("")


def _canonical_lookups(extractor: EntityExtractor) -> dict[str, dict[str, str]]:
    lookups = {"technical_skill": {}, "tool": {}, "soft_skill": {}}
    for entity in extractor.skills:
        target = lookups[entity["type"]]
        target[entity["canonical"].casefold()] = entity["canonical"]
        for alias in entity["aliases"]:
            target[str(alias).casefold()] = entity["canonical"]
    for entity in extractor.soft_skills:
        target = lookups["soft_skill"]
        target[entity["canonical"].casefold()] = entity["canonical"]
        for alias in entity["aliases"]:
            target[str(alias).casefold()] = entity["canonical"]
    return lookups


def repair_ai_annotations(
    ai_annotations: pd.DataFrame,
    primary: pd.DataFrame,
    extractor: EntityExtractor | None = None,
) -> AIRepairResult:
    """Canonicalize AI labels and restore exact primary source text and fingerprints."""
    _require_columns(ai_annotations, REQUIRED_COLUMNS | {"text_sha256"}, "AI annotation")
    _require_columns(primary, REQUIRED_COLUMNS, "primary annotation")
    if ai_annotations["source_row"].map(_source_key).duplicated().any():
        raise ValueError("AI annotations contain duplicate source_row values")
    if primary["source_row"].map(_source_key).duplicated().any():
        raise ValueError("Primary annotations contain duplicate source_row values")

    primary_by_key = {
        _source_key(row["source_row"]): row for _, row in primary.iterrows()
    }
    ai_keys = {_source_key(value) for value in ai_annotations["source_row"]}
    if ai_keys != set(primary_by_key):
        raise ValueError("AI and primary annotations must contain the same source_row set")

    engine = extractor or EntityExtractor()
    lookups = _canonical_lookups(engine)
    dropped: dict[str, Counter[str]] = {field: Counter() for field in SET_FIELDS}
    alias_mappings: dict[str, Counter[str]] = {field: Counter() for field in SET_FIELDS}
    input_counts = Counter()
    output_counts = Counter()
    source_text_repaired = 0
    stored_hash_matches_primary = 0
    invalid_scalar_counts = Counter()
    records: list[dict[str, str]] = []

    for _, row in ai_annotations.iterrows():
        source_key = _source_key(row["source_row"])
        primary_row = primary_by_key[source_key]
        primary_text = str(primary_row["text"])
        source_text_repaired += int(str(row["text"]) != primary_text)
        stored_hash_matches_primary += int(
            _clean(row["text_sha256"]) == _text_sha256(primary_text)
        )
        record = {
            "source_row": source_key,
            "text": primary_text,
            "text_sha256": _text_sha256(primary_text),
            "review_status": "ai_reviewed",
            "annotator": _clean(row["annotator"]) or "external_ai_annotator",
        }

        row_dropped = 0
        for field, entity_type in SET_FIELDS.items():
            canonical: set[str] = set()
            lookup = lookups[entity_type]
            for label in _split(row[field]):
                input_counts[field] += 1
                mapped = lookup.get(label.casefold())
                if mapped is None:
                    dropped[field][label] += 1
                    row_dropped += 1
                    continue
                canonical.add(mapped)
                if mapped != label:
                    alias_mappings[field][f"{label} -> {mapped}"] += 1
            record[field] = "|".join(sorted(canonical))
            output_counts[field] += len(canonical)

        education = _clean(row["gold_education"])
        education_values = [value for value in _split(education) if value in ALLOWED_EDUCATION]
        invalid_scalar_counts["gold_education"] += len(_split(education)) - len(education_values)
        record["gold_education"] = "|".join(sorted(set(education_values)))
        record["gold_experience_years"] = _normalized_years(row["gold_experience_years"])
        if _clean(row["gold_experience_years"]) and not record["gold_experience_years"]:
            invalid_scalar_counts["gold_experience_years"] += 1
        seniority = _clean(row["gold_seniority"])
        arrangement = _clean(row["gold_work_arrangement"])
        if seniority not in ALLOWED_SENIORITY:
            invalid_scalar_counts["gold_seniority"] += 1
            seniority = "unknown"
        if arrangement not in ALLOWED_ARRANGEMENT:
            invalid_scalar_counts["gold_work_arrangement"] += 1
            arrangement = "unknown"
        record["gold_seniority"] = seniority
        record["gold_work_arrangement"] = arrangement
        original_notes = _clean(row["notes"])
        repair_note = (
            "Canonicalized as an AI challenger against taxonomy v0.2; exact primary source "
            f"text restored; {row_dropped} unsupported labels removed."
        )
        record["notes"] = f"{original_notes} {repair_note}".strip()
        records.append(record)

    repaired = pd.DataFrame(records, columns=AI_CHALLENGER_COLUMNS)
    quality = validate_annotation_frame(repaired, engine)
    if not quality["valid"]:
        codes = ", ".join(issue["code"] for issue in quality["issues"])
        raise ValueError(f"Repaired AI annotations failed validation: {codes}")

    report = {
        "artifact_type": "canonical AI challenger annotations",
        "claim_status": "not_human_gold_not_ml_qg_2",
        "transformation_version": "ai_challenger_repair_v1",
        "rows_input": len(ai_annotations),
        "rows_output": len(repaired),
        "unique_source_rows": int(repaired["source_row"].nunique()),
        "source_text_rows_restored": source_text_repaired,
        "input_hashes_matching_primary_source": stored_hash_matches_primary,
        "review_status": "ai_reviewed",
        "annotators": dict(Counter(repaired["annotator"])),
        "field_summary": {
            field: {
                "input_labels": input_counts[field],
                "output_labels": output_counts[field],
                "dropped_labels": sum(dropped[field].values()),
                "alias_mappings": [
                    {"mapping": label, "count": count}
                    for label, count in alias_mappings[field].most_common()
                ],
                "top_dropped": [
                    {"label": label, "count": count}
                    for label, count in dropped[field].most_common(25)
                ],
            }
            for field in SET_FIELDS
        },
        "invalid_scalar_values_replaced": dict(invalid_scalar_counts),
        "annotation_quality": quality,
        "limitations": [
            "Labels were produced by an AI annotator and cannot satisfy the two-human ML-QG-2 gate.",
            "Unsupported concepts were dropped rather than silently expanding taxonomy v0.2.",
            "Agreement with primary gold is descriptive AI-vs-human evidence only.",
        ],
    }
    return AIRepairResult(repaired, report)


def evaluate_ai_challenger_agreement(
    primary: pd.DataFrame, ai_challenger: pd.DataFrame
) -> dict[str, Any]:
    """Reuse field metrics while explicitly refusing a human-agreement claim."""
    comparison = ai_challenger.copy()
    comparison["review_status"] = "reviewed"
    report = evaluate_annotation_agreement(
        primary, comparison, minimum_documents=len(comparison)
    )
    report["evaluation_type"] = "AI-vs-human annotation comparison"
    report["claim_status"] = "descriptive_only_not_human_agreement"
    report["annotators_are_independent"] = False
    report["ready_for_ml_qg_2"] = False
    report["ai_annotators"] = report.pop("secondary_annotators")
    report["warnings"].insert(
        0,
        "The challenger annotator is AI; these metrics do not satisfy ML-QG-2 regardless of score.",
    )
    return report


def write_repaired_workbook(
    annotations: pd.DataFrame, report: dict[str, Any], output: Path
) -> None:
    """Write a styled local XLSX with explicit AI-only provenance and QA summary."""
    workbook = openpyxl.Workbook()
    annotation_sheet = workbook.active
    annotation_sheet.title = "AI_Challenger_Annotations"
    annotation_sheet.append(list(annotations.columns))
    for row in annotations.itertuples(index=False, name=None):
        annotation_sheet.append(list(row))
    annotation_sheet.freeze_panes = "A2"
    annotation_sheet.auto_filter.ref = annotation_sheet.dimensions
    widths = {
        "A": 12,
        "B": 90,
        "C": 22,
        "D": 32,
        "E": 28,
        "F": 28,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 20,
        "K": 16,
        "L": 28,
        "M": 70,
    }
    for column, width in widths.items():
        annotation_sheet.column_dimensions[column].width = width
    for cell in annotation_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in annotation_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Transformation_QA")
    summary.append(["Metric", "Value"])
    summary_rows = (
        ("Artifact type", report["artifact_type"]),
        ("Claim status", report["claim_status"]),
        ("Rows", report["rows_output"]),
        ("Source texts restored", report["source_text_rows_restored"]),
        ("Input hashes matching primary", report["input_hashes_matching_primary_source"]),
        ("Output validation", report["annotation_quality"]["valid"]),
        ("Review status", report["review_status"]),
    )
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 70
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C65911")

    dropped_sheet = workbook.create_sheet("Dropped_Labels")
    dropped_sheet.append(["Field", "Label", "Count"])
    for field, field_report in report["field_summary"].items():
        for item in field_report["top_dropped"]:
            dropped_sheet.append([field, item["label"], item["count"]])
    dropped_sheet.auto_filter.ref = dropped_sheet.dimensions
    dropped_sheet.freeze_panes = "A2"
    for cell in dropped_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7F6000")

    guide = workbook.create_sheet("README")
    guide.append(["SkillPulse AI challenger — usage boundary"])
    guide.append([
        "This workbook contains canonicalized AI annotations. It is not human gold, "
        "not a blind second annotation, and cannot satisfy ML-QG-2."
    ])
    guide.append([
        "Use it only for descriptive AI-vs-human comparison, taxonomy gap analysis, "
        "and future challenger experiments."
    ])
    guide.column_dimensions["A"].width = 120
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", type=Path, required=True)
    parser.add_argument(
        "--primary", type=Path, default=Path("data/annotations/gold_sample.csv")
    )
    parser.add_argument(
        "--output-csv", type=Path, default=Path("data/annotations/ai_challenger_annotations.csv")
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("data/annotations/SkillPulse_AI_Annotation_Challenger_FIXED.xlsx"),
    )
    parser.add_argument(
        "--repair-report", type=Path, default=Path("reports/ai_challenger_repair.json")
    )
    parser.add_argument(
        "--agreement-report", type=Path, default=Path("reports/ai_challenger_agreement.json")
    )
    args = parser.parse_args()

    source = read_annotation_workbook(args.input_xlsx)
    primary = pd.read_csv(args.primary, dtype=str, keep_default_na=False)
    result = repair_ai_annotations(source, primary)
    agreement = evaluate_ai_challenger_agreement(primary, result.annotations)
    write_csv_atomic(result.annotations, args.output_csv)
    write_repaired_workbook(result.annotations, result.report, args.output_xlsx)
    args.repair_report.parent.mkdir(parents=True, exist_ok=True)
    args.repair_report.write_text(
        json.dumps(result.report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    args.agreement_report.parent.mkdir(parents=True, exist_ok=True)
    args.agreement_report.write_text(
        json.dumps(agreement, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "output_csv": str(args.output_csv),
                "output_xlsx": str(args.output_xlsx),
                "rows": len(result.annotations),
                "quality_valid": result.report["annotation_quality"]["valid"],
                "claim_status": result.report["claim_status"],
                "ready_for_ml_qg_2": agreement["ready_for_ml_qg_2"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
